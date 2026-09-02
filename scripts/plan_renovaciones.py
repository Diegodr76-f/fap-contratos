# -*- coding: utf-8 -*-
"""
Plan de renovaciones y procesos nuevos FAP 2027.

Lee la hoja "2026" del «Sistema de Alertas de Contratos FIAS» y responde, contrato
por contrato, la única pregunta que ordena todo el año: ¿se puede renovar, o hay
que hacer un nuevo proceso administrativo?

La regla es del FIAS: un contrato se renueva UNA sola vez. El que en 2026 se firmó
como renovación ya gastó su cupo y para 2027 necesita proceso nuevo. El que se
firmó como nuevo todavía puede renovarse.

La meta es tener los expedientes precontractuales hechos y revisados antes de
enero. La firma no depende de nosotros: el PAG se aprueba en promedio hasta el 15
de enero, y sin PAG no se puede suscribir ni pedir una cotización en firme, porque
es el PAG el que fija el presupuesto de cada área.

    python3 scripts/plan_renovaciones.py Sistema_Alertas_Contratos_FIAS.xlsx [carpeta_salida]

Escribe (fuera del repositorio, porque llevan datos de contratos y correos):
    Anexo_Renovaciones_2027_FAP.xlsx   maestro + una hoja por AC + calendario + simulación
    correos/<AC>.txt                   el correo de consulta ya redactado
"""
import sys, os, datetime, re, unicodedata, argparse, html, urllib.parse
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- parámetros
HOY = datetime.date(2026, 9, 1)

# Objetos que no son servicio recurrente: nunca admiten renovación y quedan
# fuera de esta campaña (van por su propio proceso, con TdR).
NO_RECURRENTES = {"Consultoría", "Adquisición de equipos de campo"}

RENOVACION = "Renovación"
NUEVO = "Nuevo proceso"

# La causal con la que sale la mayoría de los procesos nuevos: el mismo proveedor
# ya está calificado, el servicio es recurrente y en territorio son de los pocos
# que facturan y operan legalmente.
CAUSAL_DIRECTA = ("Contratación directa — proveedor calificado y recurrencia del servicio "
                  "(en territorio, de los pocos proveedores que facturan y operan legalmente)")

# Fecha en que se aprueba el PAG. Es la compuerta de la firma: antes de eso no se
# puede suscribir ni pedir cotización en firme. En promedio se aprueba hasta el 15
# de enero de cada año.
PAG = datetime.date(2027, 1, 15)
DIAS_TRAS_PAG = 10            # cotización en firme con el techo presupuestario + instrumento

# Modelo de tiempos, calibrado contra 2026: días de revisión = BASE + 0,3 por cada
# proceso en cola el día de la solicitud. Reproduce las medianas observadas
# (renovación 31 d con ~50 en cola, contratación directa 21 d con ~20).
DIAS_BASE = 15
DIAS_POR_PENDIENTE = 0.3

# Cupo semanal de ingreso de expedientes. La unidad legal firmó 5,4 instrumentos
# por semana de promedio en 2026, 8,8 sostenidos en febrero-abril y 13 en su mejor
# semana; 10 sostiene el plan sin repetir el atasco de enero (60 procesos en cola).
CUPO_SEMANAL = 10
PRIMERA_SEMANA = datetime.date(2026, 9, 21)   # lunes

# Los sucesores no arrancan todos el mismo día: cada grupo tiene su fecha tope.
CORTE_1 = datetime.date(2027, 1, 1)
CORTE_2 = datetime.date(2027, 2, 1)

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def dl(d):
    return f"{d.day} de {MESES[d.month - 1]} de {d.year}"


def plural(n, sing, plu):
    return f"{n} {sing if n == 1 else plu}"


def norm(s):
    s = "" if s is None else str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", s)).strip()


def fecha(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    return v if isinstance(v, datetime.date) else None


def grupo_firma(inicio):
    if inicio is None:
        return "—"
    if inicio <= CORTE_1:
        return "1-ene"
    return "1-feb" if inicio <= CORTE_2 else "posterior"


def leer(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["2026"]
    filas = list(ws.iter_rows(values_only=True))
    hdr = list(filas[1])
    out = [dict(zip(hdr, r)) for r in filas[2:] if dict(zip(hdr, r)).get("Nro. DE CONTRATO")]
    wb.close()
    return out


def clasificar(filas):
    """Renovación o nuevo proceso. No hay más categorías."""
    universo, fuera = [], []
    for d in filas:
        if d.get("Estado de Gestión") != "Activo":
            continue
        cat = d["CATEGORIA DEL PROCESO/ BIEN O SERVICIO"]
        recurrente = cat not in NO_RECURRENTES
        renovable = d["TIPO DE CONTRATO"] == "Nuevo" and recurrente
        fin = fecha(d["FECHA DE FINALIZACIÓN"])
        inicio27 = fin + datetime.timedelta(days=1) if fin else None
        reg = {
            "nro": d["Nro. DE CONTRATO"],
            "detalle": str(d["DETALLE DEL CONTRATO"] or "").strip(),
            "area": d["Área Protegida"],
            "categoria": cat,
            "monto": float(d["Monto (incluido IVA)"] or 0),
            # «Valor o plazo total» trae el monto ya sumadas las adendas de valor.
            # En las adendas de plazo repite el monto original, así que nunca queda por debajo.
            "monto_total": float(d["Valor o plazo total "] or d["Monto (incluido IVA)"] or 0),
            "fin": fin,
            "inicio27": inicio27,
            "grupo": grupo_firma(inicio27),
            "tipo2026": d["TIPO DE CONTRATO"],
            "proveedor": str(d["Nombre del Proveedor "] or "").strip(),
            "ac": d["Administrador/a de contrato"],
            "correo": d["Correo electrónico AC"],
            "adenda": d["Tiene adenda"],
            "tipo": RENOVACION if renovable else NUEVO,
            "modalidad": "" if renovable else CAUSAL_DIRECTA,
        }
        (universo if recurrente else fuera).append(reg)
    return universo, fuera


def programar(universo):
    """Reparte el universo en semanas de ingreso a la Unidad Operativa.

    Primero lo que arranca antes; dentro de eso, los procesos nuevos antes que
    las renovaciones, porque llevan un documento más. Y dentro de cada nivel,
    agrupado por categoría: revisar diez expedientes de combustible seguidos
    cuesta menos que alternarlos con mantenimientos y arriendos, porque el
    revisor contrasta contra el mismo modelo en vez de cambiar de marco.

    Entrar temprano no adelanta la firma más allá del PAG. Lo que define es la
    POSICIÓN EN LA COLA de firma del 15 de enero en adelante.
    """
    orden = sorted(universo, key=lambda x: (x["inicio27"], 0 if x["tipo"] == NUEVO else 1,
                                            x["categoria"], -x["monto"]))
    for i, x in enumerate(orden):
        x["semana"] = PRIMERA_SEMANA + datetime.timedelta(weeks=i // CUPO_SEMANAL)
        x["docs_listos"] = x["semana"] - datetime.timedelta(days=7)
    return orden


# ------------------------------------------------------------------ simulación
def preparar(procesos):
    """Cuándo queda revisado cada expediente: solicitud + base + carga de la cola."""
    ps = sorted(procesos, key=lambda x: x["semana"])
    for i, x in enumerate(ps):
        cola = sum(1 for y in ps[:i] if y.get("revisado") and y["revisado"] > x["semana"])
        x["revisado"] = x["semana"] + datetime.timedelta(days=round(DIAS_BASE + DIAS_POR_PENDIENTE * cola))
    return ps


def firmar(procesos, capacidad, pag=PAG, prioridad="arranque"):
    """Reparte las firmas desde que el PAG las habilita, a `capacidad` por semana.

    `prioridad` decide a quién se firma primero cuando no alcanza para todos:
      · "arranque"  — por fecha de arranque del sucesor; minimiza los días totales
                      de servicio sin contrato.
      · "a_tiempo"  — primero los que todavía pueden firmarse antes de arrancar;
                      minimiza el NÚMERO de contratos que quedan irregulares.
    """
    claves = {
        "arranque": lambda x: (x["inicio27"], x["revisado"]),
        "a_tiempo": lambda x: (0 if x["inicio27"] > CORTE_1 else 1, x["inicio27"], x["revisado"]),
    }
    arranque = pag + datetime.timedelta(days=DIAS_TRAS_PAG)
    for i, x in enumerate(sorted(procesos, key=claves[prioridad])):
        x["firma"] = max(x["revisado"], arranque) + datetime.timedelta(weeks=i // capacidad)
        x["retro"] = max(0, (x["firma"] - x["inicio27"]).days)
    return procesos


def resumen_simulacion(procesos):
    r = sorted(x["retro"] for x in procesos)
    g2 = [x for x in procesos if x["grupo"] == "1-feb"]
    return {
        "mediana": r[len(r) // 2] if r else 0,
        "maxima": r[-1] if r else 0,
        "total": sum(r),
        "sin_retro": sum(1 for v in r if v == 0),
        "g2_ok": sum(1 for x in g2 if x["firma"] <= datetime.date(2027, 1, 31)),
        "g2": len(g2),
        "ultima": max(x["firma"] for x in procesos),
    }


def simular(universo, capacidad, pag=PAG, prioridad="arranque"):
    ps = [dict(x) for x in universo]
    return resumen_simulacion(firmar(preparar(ps), capacidad, pag, prioridad))


def simular_sin_plan(universo, capacidad, pag=PAG):
    """El contrafactual: las solicitudes vuelven a llegar todas en enero."""
    ps = []
    for i, x in enumerate(universo):
        y = dict(x)
        y["semana"] = datetime.date(2027, 1, 4) + datetime.timedelta(weeks=i % 4)
        ps.append(y)
    return resumen_simulacion(firmar(preparar(ps), capacidad, pag))



# ------------------------------------------------- formulario de confirmación
# Las respuestas de las administradoras se recogen con Microsoft Forms, que está
# en el plan básico y no necesita conectores premium. Cada contrato lleva su
# propio enlace con los datos ya rellenados, así la AC nunca escribe un número de
# contrato y las respuestas se pueden cruzar sin ambigüedad. El paso a paso para
# crear el formulario está en plan/FORMULARIO_CONFIRMACION.md.
#
# En el enlace de ejemplo que entrega Forms («obtener vínculo para rellenar
# previamente las respuestas»), estas palabras se escriben tal cual en cada
# campo; el script las reemplaza por los datos de cada contrato.
TOKENS_FORM = {
    "NROCONTRATO": "nro",
    "AREAPROTEGIDA": "area",
    "DETALLESERVICIO": "detalle",
    "ADMINISTRADORA": "ac",
    "TIPO2027": "tipo",
}


def enlace_form(plantillas, x):
    """Enlace de Forms con los datos del contrato ya rellenados.

    `plantillas` es un diccionario {tipo: url}. Son dos formularios distintos
    porque la pregunta que sigue no es la misma: al renovable se le pregunta si
    renueva o va a proceso nuevo, y al de proceso nuevo si va por contratación
    directa o por comparación de precios. La ramificación de Forms solo funciona
    sobre preguntas de opción del propio formulario, así que separarlos es lo que
    permite que cada AC vea únicamente la pregunta que le toca."""
    plantilla = (plantillas or {}).get(x["tipo"], "")
    if not plantilla:
        return ""
    url = plantilla
    for token, campo in TOKENS_FORM.items():
        url = url.replace(token, urllib.parse.quote(str(x.get(campo) or ""), safe=""))
    return url


# Cómo se reconocen las columnas del Excel de respuestas: por un trozo del texto
# de la pregunta, para que sobreviva a que alguien reescriba el enunciado.
COLUMNAS_RESPUESTA = {
    "nro": ["número de contrato", "numero de contrato", "n.º de contrato"],
    "continua": ["mantener este servicio"],
    "via_elegida": ["se tramita"],
    "proveedor": ["mismo proveedor"],
    "consumo": ["consumo ejecutado"],
    "monto27": ["monto estimado"],
    "clausula": ["cláusula de renovación", "clausula de renovacion"],
    "obs": ["observaciones"],
}


def leer_respuestas(xlsx):
    """Lee el Excel que alimenta Forms y devuelve {n.º de contrato: respuesta}.

    Se queda con la última respuesta de cada contrato: si una AC corrige, la
    corrección manda."""
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        return {}
    hdr = [norm(c) for c in filas[0]]
    idx = {}
    for campo, pistas in COLUMNAS_RESPUESTA.items():
        for i, h in enumerate(hdr):
            if any(norm(p) in h for p in pistas):
                idx[campo] = i
                break
    if "nro" not in idx:
        sys.exit("ERROR: el Excel de respuestas no tiene una columna con el número de contrato.")

    out = {}
    for r in filas[1:]:
        nro = r[idx["nro"]]
        if not nro:
            continue
        out[str(nro).strip().upper()] = {
            campo: (r[i] if i < len(r) else None) for campo, i in idx.items()
        }
    return out


def cruzar(universo, respuestas):
    """Marca cada contrato con lo que respondió su administradora."""
    for x in universo:
        r = respuestas.get(str(x["nro"]).strip().upper())
        x["respondido"] = "Sí" if r else "Sin responder"
        x["continua"] = (r or {}).get("continua") or ""
        x["via_elegida"] = (r or {}).get("via_elegida") or ""
        x["consumo26"] = (r or {}).get("consumo")
        x["monto27"] = (r or {}).get("monto27")
        x["clausula"] = (r or {}).get("clausula") or ""
        x["observaciones"] = (r or {}).get("obs") or ""
    return universo


def informe_respuestas(universo):
    """Quién respondió, quién no, y qué se cae del plan."""
    porac = defaultdict(list)
    for x in universo:
        porac[x["ac"]].append(x)
    filas = []
    for ac, g in sorted(porac.items()):
        resp = [y for y in g if y["respondido"] == "Sí"]
        baja = [y for y in resp if norm(y["continua"]).startswith("no")]
        filas.append({
            "ac": ac, "correo": g[0]["correo"], "total": len(g),
            "respondidos": len(resp), "pendientes": len(g) - len(resp),
            "no_continuan": len(baja),
            "monto_confirmado": sum(float(y["monto27"] or y["monto"] or 0)
                                    for y in resp if not norm(y["continua"]).startswith("no")),
        })
    return filas


# ------------------------------------------------------------------- salidas
AZUL = "00249C"
COL = {RENOVACION: "E4F2E9", NUEVO: "FAF0DE"}

CAMPOS = [
    ("nro", "N.º de contrato 2026", 21),
    ("detalle", "Detalle", 46),
    ("area", "Área protegida", 34),
    ("categoria", "Categoría", 16),
    ("monto", "Monto 2026 (USD)", 15),
    ("monto_total", "Monto total con adendas (USD)", 24),
    ("tipo2026", "Tipo 2026", 12),
    ("tipo", "Para 2027", 15),
    ("modalidad", "Modalidad y causal", 46),
    ("fin", "Vence", 12),
    ("inicio27", "Arranca el sucesor", 18),
    ("grupo", "Grupo de firma", 14),
    ("semana", "Semana de ingreso", 17),
    ("docs_listos", "Documentos de la AC listos", 24),
    ("proveedor", "Proveedor 2026", 34),
    ("ac", "Administrador/a", 20),
]

# Columnas que solo aparecen cuando se cruzó el Excel de respuestas de Forms.
CAMPOS_RESPUESTA = [
    ("respondido", "¿Respondió?", 13),
    ("continua", "¿Continúa en 2027?", 26),
    ("via_elegida", "Vía que eligió la AC", 40),
    ("consumo26", "Consumo ejecutado 2026", 22),
    ("monto27", "Monto estimado 2027", 20),
    ("clausula", "¿Tiene cláusula de renovación?", 28),
    ("observaciones", "Observaciones de la AC", 46),
]


def campos(universo):
    """El maestro crece con las respuestas solo si ya las hay."""
    hay = any("respondido" in x for x in universo)
    return CAMPOS + (CAMPOS_RESPUESTA if hay else [])


def _cab(ws, fila, titulos, anchos):
    for i, (t, a) in enumerate(zip(titulos, anchos), start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


def _fila(ws, r, x, cols=None):
    for i, (k, _, _) in enumerate(cols or CAMPOS, start=1):
        v = x.get(k)
        if isinstance(v, datetime.date):
            v = v.isoformat()
        c = ws.cell(row=r, column=i, value=v)
        c.alignment = Alignment(vertical="top", wrap_text=(i in (2, 3, 8, 14)))
        if k in ("monto", "monto_total"):
            c.number_format = '#,##0.00'
        c.fill = PatternFill("solid", fgColor=COL[x["tipo"]])


def escribir_xlsx(universo, fuera, destino):
    wb = openpyxl.Workbook()

    # --- Resumen
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "PLAN DE RENOVACIONES Y PROCESOS NUEVOS · FAP 2027"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws["A2"] = f"Corte {HOY.isoformat()} · fuente: Sistema de Alertas de Contratos FIAS, hoja 2026"
    ws["A2"].font = Font(size=10, color="6B7180")
    tc = Counter(x["tipo"] for x in universo)
    monto = defaultdict(float)
    for x in universo:
        monto[x["tipo"]] += x["monto"]
    gc = Counter(x["grupo"] for x in universo)
    gm = defaultdict(float)
    for x in universo:
        gm[x["grupo"]] += x["monto"]
    filas = [
        ("Contratos activos de servicios recurrentes en áreas protegidas", len(universo),
         sum(x["monto"] for x in universo)),
        ("SE PUEDEN RENOVAR — el contrato 2026 se firmó como nuevo", tc[RENOVACION], monto[RENOVACION]),
        ("NUEVO PROCESO ADMINISTRATIVO — ya renovaron en 2026", tc[NUEVO], monto[NUEVO]),
        ("", None, None),
        ("Sucesor arranca el 1 de enero de 2027", gc["1-ene"], gm["1-ene"]),
        ("Sucesor arranca el 1 de febrero de 2027", gc["1-feb"], gm["1-feb"]),
        ("Sucesor arranca después de febrero de 2027", gc["posterior"], gm["posterior"]),
        ("", None, None),
        ("Fuera de esta campaña: objeto no recurrente", len(fuera), sum(x["monto"] for x in fuera)),
    ]
    _cab(ws, 4, ["Concepto", "N.º", "Monto 2026 (USD)"], [62, 10, 20])
    for i, (t, n, m) in enumerate(filas, start=5):
        ws.cell(row=i, column=1, value=t)
        if n is not None:
            ws.cell(row=i, column=2, value=n)
            c = ws.cell(row=i, column=3, value=round(m, 2))
            c.number_format = '#,##0.00'

    # --- Maestro
    cols = campos(universo)
    ws = wb.create_sheet("Maestro")
    _cab(ws, 1, [t for _, t, _ in cols], [a for _, _, a in cols])
    for i, x in enumerate(sorted(universo, key=lambda y: (y["semana"], y["tipo"], y["nro"])), start=2):
        _fila(ws, i, x, cols)

    # --- Calendario
    ws = wb.create_sheet("Calendario")
    _cab(ws, 1, ["Semana de ingreso", "Expedientes", "Renovaciones", "Procesos nuevos",
                 "Arrancan el", "Categorías del lote"], [18, 12, 14, 16, 14, 46])
    sem = defaultdict(list)
    for x in universo:
        sem[x["semana"]].append(x)
    for i, s in enumerate(sorted(sem), start=2):
        g = sem[s]
        c = Counter(y["tipo"] for y in g)
        arr = sorted({y["inicio27"] for y in g})
        cats = ", ".join(f"{k} ({v})" for k, v in Counter(y["categoria"] for y in g).most_common())
        vals = [s.isoformat(), len(g), c[RENOVACION], c[NUEVO],
                " / ".join(a.isoformat() for a in arr[:2]) + ("…" if len(arr) > 2 else ""), cats]
        for j, v in enumerate(vals, start=1):
            cel = ws.cell(row=i, column=j, value=v)
            cel.alignment = Alignment(vertical="top", wrap_text=(j == 6))

    # --- Carga por AC
    ws = wb.create_sheet("Carga por AC")
    _cab(ws, 1, ["Administrador/a", "Correo", "Total", "Se renuevan", "Proceso nuevo",
                 "Monto 2026 (USD)", "Primera semana", "Responde antes de"],
         [24, 26, 8, 13, 15, 18, 16, 18])
    porac = defaultdict(list)
    for x in universo:
        porac[x["ac"]].append(x)
    for i, (ac, g) in enumerate(sorted(porac.items(), key=lambda kv: min(y["semana"] for y in kv[1])), start=2):
        c = Counter(y["tipo"] for y in g)
        s = min(y["semana"] for y in g)
        vals = [ac, g[0]["correo"], len(g), c[RENOVACION], c[NUEVO],
                round(sum(y["monto"] for y in g), 2), s.isoformat(),
                (s - datetime.timedelta(days=5)).isoformat()]
        for j, v in enumerate(vals, start=1):
            cel = ws.cell(row=i, column=j, value=v)
            if j == 6:
                cel.number_format = '#,##0.00'

    # --- Simulación de firma
    ws = wb.create_sheet("Simulación firma")
    ws["A1"] = "¿Cuándo saldría firmado cada contrato?"
    ws["A1"].font = Font(bold=True, size=12, color=AZUL)
    ws["A2"] = (f"El PAG se aprueba en promedio hasta el {dl(PAG)}. Antes de esa fecha no se puede "
                "suscribir ni pedir cotización en firme, así que la retroactividad no se elimina: "
                "se reduce. Tener el expediente listo mueve la firma de mayo-junio a marzo-abril.")
    ws["A2"].font = Font(size=10, color="6B7180")
    _cab(ws, 4, ["Escenario", "Firmas por semana", "Prioridad de firma",
                 "Retroactividad mediana", "Máxima", "Días-contrato totales",
                 "Contratos sin retroactividad", "Última firma"],
         [40, 18, 30, 22, 12, 22, 26, 14])
    r = 5
    etiquetas = {"arranque": "Por arranque del sucesor", "a_tiempo": "Primero los que llegan a tiempo"}
    for cap in (9, 13):
        for pr in ("arranque", "a_tiempo"):
            d = simular(universo, cap, prioridad=pr)
            for j, v in enumerate(["Con el plan: expedientes listos en diciembre", cap, etiquetas[pr],
                                   d["mediana"], d["maxima"], d["total"],
                                   f"{d['sin_retro']} de {len(universo)}", d["ultima"].isoformat()], start=1):
                ws.cell(row=r, column=j, value=v)
            r += 1
    for cap in (9, 13):
        d = simular_sin_plan(universo, cap)
        for j, v in enumerate(["Sin plan: las solicitudes llegan en enero, como 2026", cap,
                               etiquetas["arranque"], d["mediana"], d["maxima"], d["total"],
                               f"{d['sin_retro']} de {len(universo)}", d["ultima"].isoformat()], start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(italic=True)
        r += 1
    c = ws.cell(row=r + 1, column=1,
                value="Referencia 2026 real: 64 días de retroactividad mediana, 151 de máxima, en el 100 % de las renovaciones.")
    c.font = Font(italic=True, color="6B7180")

    # --- Una hoja por AC
    for ac, g in sorted(porac.items()):
        ws = wb.create_sheet(re.sub(r"[\\/*?:\[\]]", "", str(ac))[:31])
        c = Counter(y["tipo"] for y in g)
        ws["A1"] = f"{ac} · {len(g)} contratos · {c[RENOVACION]} se renuevan · {c[NUEVO]} van por proceso nuevo"
        ws["A1"].font = Font(bold=True, size=12, color=AZUL)
        _cab(ws, 3, [t for _, t, _ in cols], [a for _, _, a in cols])
        for i, x in enumerate(sorted(g, key=lambda y: (y["tipo"], y["semana"])), start=4):
            _fila(ws, i, x, cols)

    # --- Respuestas de las administradoras
    if any("respondido" in x for x in universo):
        ws = wb.create_sheet("Respuestas")
        ws["A1"] = "Confirmación de las administradoras"
        ws["A1"].font = Font(bold=True, size=12, color=AZUL)
        ws["A2"] = ("Leído del Excel que alimenta Microsoft Forms. Cada contrato tiene su enlace "
                    "propio, así que las respuestas se cruzan por número de contrato, sin ambigüedad.")
        ws["A2"].font = Font(size=10, color="6B7180")
        _cab(ws, 4, ["Administrador/a", "Correo", "Contratos", "Respondidos", "Pendientes",
                     "No continúan", "Monto confirmado 2027 (USD)"],
             [24, 26, 11, 12, 12, 13, 26])
        for i, f in enumerate(informe_respuestas(universo), start=5):
            vals = [f["ac"], f["correo"], f["total"], f["respondidos"], f["pendientes"],
                    f["no_continuan"], round(f["monto_confirmado"], 2)]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=i, column=j, value=v)
                if j == 7:
                    c.number_format = '#,##0.00'
                if j == 5 and v:
                    c.font = Font(bold=True, color="A33A3A")

    # --- Fuera de la campaña
    ws = wb.create_sheet("Fuera de la campaña")
    ws["A1"] = "Objetos no recurrentes: no admiten renovación y van por su propio proceso"
    ws["A1"].font = Font(bold=True, size=11, color=AZUL)
    _cab(ws, 3, [t for _, t, _ in CAMPOS[:9]], [a for _, _, a in CAMPOS[:9]])
    for i, x in enumerate(fuera, start=4):
        for j, (k, _, _) in enumerate(CAMPOS[:9], start=1):
            v = x.get(k)
            ws.cell(row=i, column=j, value=v.isoformat() if isinstance(v, datetime.date) else v)

    wb.save(destino)
    return destino


CUERPO = """Asunto: Contratos 2027 de {area_corta} — cuáles se renuevan y cuáles no

Estimada/o {ac}:

Estamos armando el plan de contratación 2027 del FAP. La regla del FIAS es que un
contrato se renueva UNA sola vez: no hay renovación sobre renovación. Revisé tus
{n} contratos vigentes y este es el resultado.

{listas}
La meta es llegar a fin de año con todos los expedientes hechos y revisados. La
firma es otra cosa: el PAG se aprueba en promedio hasta el 15 de enero, y sin PAG
no se puede suscribir ni pedir una cotización en firme, porque es el PAG el que
fija tu presupuesto. Lo que sí depende de nosotros es que el día que salga no
quede ningún documento pendiente — y que tu expediente esté adelante en la fila,
no atrás.

Te pido una cosa, antes del {responde}: confirmar contrato por contrato si el área
necesita mantener el servicio en 2027. Si alguno ya no se requiere, decirlo
también: es la única forma de no tramitar lo que no se va a usar.

AHORA, sin esperar al PAG:

{bloque1}

CUANDO EL PAG ESTÉ APROBADO — no lo prepares todavía, el monto sale de ahí:

{bloque2}

Sobre el monto: cuando toque cotizar, toma como base el consumo ejecutado de este
año, no el presupuesto del contrato vigente. En 2026, 11 de las 13 adendas fueron
aumentos de valor por consumo subestimado.

Tu primer lote entra la semana del {semana} y tus documentos deben estar listos el
{docs}. Esa semana la Unidad Operativa recibe hasta {cupo} expedientes de todo el FAP,
así que llegar tarde a tu semana significa esperar a la siguiente.

Todo se envía por el formulario de procesos administrativos. Los expedientes que
lleguen por correo no entran a la cola: en 2026 hubo que reconstruir 18 contratos
que ingresaron por fuera del formulario.

En La Mágica ya tienes las plantillas de los documentos precontractuales.

Gracias,
{firma}
"""

ENCABEZADOS = {
    RENOVACION: ("SE PUEDEN RENOVAR ({n})",
                 "El contrato de 2026 se firmó como nuevo, así que conserva su única renovación."),
    NUEVO: ("NECESITAN UN NUEVO PROCESO ADMINISTRATIVO ({n})",
            "El contrato de 2026 ya era una renovación: el cupo está agotado. Salen por "
            "contratación directa con el criterio de proveedor calificado y recurrencia del "
            "servicio — el mismo proveedor ya está evaluado, el servicio es continuo y en "
            "territorio son de los pocos que facturan y operan legalmente. Eso hay que "
            "sostenerlo en el informe de justificación."),
}

BLOQUE_1 = {
    RENOVACION: ("Para los {n} que se renuevan: informe de satisfacción firmado por ti y por el "
                 "responsable del área, con el análisis técnico, geográfico y económico que "
                 "justifica seguir con el mismo proveedor. Avísame también si alguno de esos "
                 "contratos NO tiene cláusula de renovación: en ese caso pasa a proceso nuevo."),
    NUEVO: ("Para los {n} de proceso nuevo: solicitud de inicio del responsable del área e "
            "informe de justificación que motive la contratación directa — proveedor calificado, "
            "recurrencia del servicio y las condiciones del territorio."),
}

BLOQUE_2 = {
    RENOVACION: ("Para los {n} que se renuevan: solicitud de cotización al proveedor para el nuevo "
                 "período con el presupuesto del PAG, su cotización y la notificación. El contrato "
                 "de renovación lo elabora la Unidad Operativa."),
    NUEVO: ("Para los {n} de proceso nuevo: solicitud de cotización en firme con el presupuesto "
            "asignado, cotización del proveedor, y la orden o notificación que corresponda."),
}


def escribir_correos(universo, carpeta, plantillas_form=None, firma="Unidad Legal · FAP"):
    """Un correo por administradora, en texto y en HTML.

    El HTML es el que conviene enviar: lleva el botón «Confirmar» de cada
    contrato, que abre el formulario con los datos ya rellenados. Se abre en el
    navegador, se copia con Ctrl+A / Ctrl+C y se pega en Outlook."""
    os.makedirs(carpeta, exist_ok=True)
    porac = defaultdict(list)
    for x in universo:
        porac[x["ac"]].append(x)

    for ac, g in porac.items():
        primera = min(y["semana"] for y in g)
        responde = primera - datetime.timedelta(days=5)
        docs = min(y["docs_listos"] for y in g)
        areas = sorted({y["area"] for y in g})
        area_corta = areas[0] if len(areas) == 1 else f"tus {len(areas)} áreas protegidas"
        c = Counter(y["tipo"] for y in g)

        listas, listas_html = [], []
        for tipo in (RENOVACION, NUEVO):
            v = sorted([y for y in g if y["tipo"] == tipo], key=lambda y: y["semana"])
            if not v:
                continue
            titulo, nota = ENCABEZADOS[tipo]
            filas, filas_html = [], []
            for y in v:
                enlace = enlace_form(plantillas_form, y)
                det = (y["detalle"][:52] + "..") if len(y["detalle"]) > 54 else y["detalle"]
                extra = " (incluye adenda)" if y["monto_total"] > y["monto"] + 0.5 else ""
                fila = (f"  · {y['nro']}  {det}\n"
                        f"      {y['area']}\n"
                        f"      USD {y['monto_total']:,.2f} este año{extra} · {y['proveedor']}\n"
                        f"      vence {dl(y['fin'])} · el sucesor arranca el {dl(y['inicio27'])}"
                        f" · expediente la semana del {dl(y['semana'])}")
                if enlace:
                    fila += f"\n      Confirmar: {enlace}"
                filas.append(fila)
                filas_html.append(_fila_html(y, enlace, tipo))
            listas.append(titulo.format(n=len(v)) + "\n" + nota + "\n\n" + "\n".join(filas) + "\n")
            listas_html.append(_lista_html(titulo.format(n=len(v)), nota, filas_html, tipo))

        bloques = []
        for fuente in (BLOQUE_1, BLOQUE_2):
            bloques.append("\n\n".join(
                fuente[t].format(n=plural(c[t], "contrato", "contratos"))
                for t in (RENOVACION, NUEVO) if c[t]))

        datos = dict(area_corta=area_corta, ac=ac, n=len(g), responde=dl(responde),
                     bloque1=bloques[0], bloque2=bloques[1], semana=dl(primera),
                     docs=dl(docs), cupo=CUPO_SEMANAL, firma=firma)
        base = re.sub(r"[^\w]+", "_", str(ac))

        with open(os.path.join(carpeta, base + ".txt"), "w", encoding="utf-8") as f:
            f.write(CUERPO.format(listas="\n".join(listas), **datos))
        with open(os.path.join(carpeta, base + ".html"), "w", encoding="utf-8") as f:
            f.write(_correo_html(listas_html, bool(plantillas_form), **datos))
    return len(porac)


# ---- versión HTML del correo, para pegar en Outlook con los enlaces vivos ----
COLOR = {RENOVACION: ("#14603D", "#E4F2E9"), NUEVO: ("#8A5210", "#FAF0DE")}


def _fila_html(y, enlace, tipo):
    tinta, fondo = COLOR[tipo]
    boton = ("" if not enlace else
             f'<a href="{html.escape(enlace, quote=True)}" '
             f'style="display:inline-block;background:#00249C;color:#ffffff;text-decoration:none;'
             f'font-weight:700;font-size:13px;padding:8px 14px;border-radius:7px;white-space:nowrap">'
             f'Confirmar</a>')
    return (
        f'<tr>'
        f'<td style="padding:12px 14px;border-bottom:1px solid #E2E6F2;vertical-align:top">'
        f'<div style="font-weight:700;color:#0B153F;font-size:14px">{html.escape(y["nro"])}</div>'
        f'<div style="color:#252A3D;font-size:14px;margin-top:2px">{html.escape(y["detalle"])}</div>'
        f'<div style="color:#5A6072;font-size:13px;margin-top:4px">{html.escape(str(y["area"]))}</div>'
        f'<div style="color:#5A6072;font-size:13px;margin-top:4px">'
        f'Vence el {dl(y["fin"])} · el sucesor arranca el {dl(y["inicio27"])} · '
        f'expediente la semana del {dl(y["semana"])}</div>'
        f'</td>'
        f'<td style="padding:12px 14px;border-bottom:1px solid #E2E6F2;vertical-align:top;'
        f'text-align:right">{boton}</td>'
        f'</tr>')


def _lista_html(titulo, nota, filas, tipo):
    tinta, fondo = COLOR[tipo]
    return (
        f'<div style="margin:26px 0 0">'
        f'<div style="background:{fondo};color:{tinta};font-weight:700;font-size:13px;'
        f'letter-spacing:.06em;text-transform:uppercase;padding:10px 14px;border-radius:8px 8px 0 0">'
        f'{html.escape(titulo)}</div>'
        f'<div style="border:1px solid #E2E6F2;border-top:none;border-radius:0 0 8px 8px">'
        f'<div style="padding:12px 14px;color:#5A6072;font-size:13.5px;line-height:1.5;'
        f'border-bottom:1px solid #E2E6F2">{html.escape(nota)}</div>'
        f'<table role="presentation" cellpadding="0" cellspacing="0" width="100%" '
        f'style="border-collapse:collapse">{"".join(filas)}</table>'
        f'</div></div>')


def _correo_html(listas_html, con_form, ac, n, responde, bloque1, bloque2,
                 semana, docs, cupo, firma, area_corta):
    def parrafo(t, color="#252A3D"):
        return (f'<p style="margin:0 0 14px;color:{color};font-size:15px;line-height:1.6">'
                f'{t}</p>')
    pedido = ('pulsar <b>Confirmar</b> en cada contrato y responder las cinco preguntas del '
              'formulario: si el servicio continúa en 2027, si sigue el mismo proveedor y '
              'cuánto se consumió realmente este año. Toma menos de un minuto por contrato.'
              if con_form else
              'confirmar contrato por contrato si el área necesita mantener el servicio en 2027.')
    bloques = "".join(
        f'<div style="margin:10px 0 0;padding:14px 16px;background:#F5F6FB;border-radius:8px;'
        f'color:#252A3D;font-size:14.5px;line-height:1.6">{html.escape(b).replace(chr(10)*2, "<br><br>")}</div>'
        for b in (bloque1, bloque2) if b)
    return (
        '<div style="font-family:Segoe UI,Calibri,Arial,sans-serif;max-width:720px;color:#252A3D">'
        + parrafo(f'Estimada/o {html.escape(str(ac))}:')
        + parrafo('Estamos armando el plan de contratación 2027 del FAP. La regla del FIAS es que '
                  'un contrato se renueva <b>una sola vez</b>: no hay renovación sobre renovación. '
                  f'Revisé tus {n} contratos vigentes y este es el resultado.')
        + "".join(listas_html)
        + '<div style="height:26px"></div>'
        + parrafo('La meta es llegar a fin de año con todos los expedientes hechos y revisados. La '
                  'firma es otra cosa: el PAG se aprueba en promedio hasta el 15 de enero, y sin PAG '
                  'no se puede suscribir ni pedir una cotización en firme, porque es el PAG el que '
                  'fija tu presupuesto. Lo que sí depende de nosotros es que el día que salga no '
                  'quede ningún documento pendiente — y que tu expediente esté adelante en la fila, '
                  'no atrás.')
        + parrafo(f'<b>Antes del {html.escape(responde)}</b>, te pido {pedido}')
        + parrafo('<b>Ahora, sin esperar al PAG:</b>') + bloques
        + '<div style="height:16px"></div>'
        + parrafo(f'Tu primer lote entra la semana del {html.escape(semana)} y tus documentos deben '
                  f'estar listos el {html.escape(docs)}. Esa semana la Unidad Operativa recibe hasta '
                  f'{cupo} expedientes de todo el FAP, así que llegar tarde a tu semana significa '
                  'esperar a la siguiente.')
        + parrafo('El expediente se envía por el formulario de procesos administrativos, como '
                  'siempre. Los que llegan por correo no entran a la cola.', "#5A6072")
        + parrafo(f'Gracias,<br>{html.escape(firma)}')
        + '</div>')


def main():
    ap = argparse.ArgumentParser(
        description="Plan de renovaciones y procesos nuevos FAP 2027.",
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    ap.add_argument("alertas", help="Sistema_Alertas_Contratos_FIAS.xlsx")
    ap.add_argument("salida", nargs="?", default=".", help="carpeta donde escribir")
    ap.add_argument("--form-renovacion", default="", metavar="URL",
                    help="enlace de pre-relleno del formulario de RENOVACIONES")
    ap.add_argument("--form-nuevo", default="", metavar="URL",
                    help="enlace de pre-relleno del formulario de PROCESOS NUEVOS. Los dos llevan "
                         "las palabras NROCONTRATO, AREAPROTEGIDA, DETALLESERVICIO, ADMINISTRADORA "
                         "y TIPO2027 escritas en sus campos; ver plan/FORMULARIO_CONFIRMACION.md")
    ap.add_argument("--respuestas", default=[], nargs="+", metavar="XLSX",
                    help="Excel(es) de respuestas, uno por formulario, para cruzarlos con el plan")
    a = ap.parse_args()
    os.makedirs(a.salida, exist_ok=True)

    universo, fuera = clasificar(leer(a.alertas))
    programar(universo)

    respuestas = {}
    for ruta in a.respuestas:
        respuestas.update(leer_respuestas(ruta))
    if respuestas:
        cruzar(universo, respuestas)

    anexo = escribir_xlsx(universo, fuera, os.path.join(a.salida, "Anexo_Renovaciones_2027_FAP.xlsx"))
    formularios = {RENOVACION: a.form_renovacion, NUEVO: a.form_nuevo}
    n = escribir_correos(universo, os.path.join(a.salida, "correos"), formularios)

    tc = Counter(x["tipo"] for x in universo)
    gc = Counter(x["grupo"] for x in universo)
    print(f"universo: {len(universo)} contratos · se renuevan {tc[RENOVACION]} · "
          f"proceso nuevo {tc[NUEVO]}")
    print(f"fuera de la campaña (objeto no recurrente): {len(fuera)}")
    print(f"arranque del sucesor: 1-ene {gc['1-ene']} · 1-feb {gc['1-feb']} · posterior {gc['posterior']}")
    print(f"ingreso: {len(set(x['semana'] for x in universo))} semanas · cupo {CUPO_SEMANAL}/semana")
    print(f"simulación de firma con el PAG el {PAG.isoformat()}:")
    for cap in (9, 13):
        p1 = simular(universo, cap)
        p0 = simular_sin_plan(universo, cap)
        print(f"   {cap} firmas/sem · con plan: mediana {p1['mediana']} d, última firma {p1['ultima']}"
              f"  ·  sin plan: mediana {p0['mediana']} d, última firma {p0['ultima']}")

    con = sum(1 for x in universo if enlace_form(formularios, x))
    if con:
        print(f"correos con enlace de confirmación: {con} de {len(universo)} contratos")
    else:
        print("correos SIN enlace de confirmación "
              "(pásale --form-renovacion y --form-nuevo para incluirlos)")

    if respuestas:
        resp = sum(1 for x in universo if x["respondido"] == "Sí")
        baja = sum(1 for x in universo if norm(x["continua"]).startswith("no"))
        print(f"respuestas: {resp} de {len(universo)} contratos confirmados · "
              f"{baja} no continúan en 2027")
        pend = [f for f in informe_respuestas(universo) if f["pendientes"]]
        if pend:
            print("pendientes por administradora:")
            for f in sorted(pend, key=lambda z: -z["pendientes"]):
                print(f"   {f['ac']:<22} {f['pendientes']:>2} de {f['total']}")

    print(f"anexo: {anexo}")
    print(f"correos redactados: {n} (.txt y .html)")


if __name__ == "__main__":
    main()
